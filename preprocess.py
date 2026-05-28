import openpyxl
import gzip
import os
import struct

def preprocess():
    xlsx_path = "SUBTLEX-NL with pos and Zipf.xlsx"
    out_dir = "src"
    
    print(f"Reading {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    sheet = wb.active
    
    data = []
    total_rows = sheet.max_row
    
    print(f"Parsing sheet rows (total rows: {total_rows})...")
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        word = row[0]
        freq = row[1]
        
        # Word and FREQcount must exist and be valid
        if word is not None and freq is not None:
            try:
                word_str = str(word).strip()
                freq_val = int(freq)
                if word_str and freq_val >= 0:
                    data.append((word_str, freq_val))
            except ValueError:
                pass
                
        if i % 50000 == 0:
            print(f"Processed {i} rows...")
            
    print(f"Found {len(data)} valid entries. Sorting by frequency descending...")
    data.sort(key=lambda x: x[1], reverse=True)
    
    os.makedirs(out_dir, exist_ok=True)
    
    words_path = os.path.join(out_dir, "words.txt")
    freqs_path = os.path.join(out_dir, "freqs.bin")
    
    print(f"Writing to {words_path} and {freqs_path}...")
    with open(words_path, "w", encoding="utf-8") as wf:
        for word, _ in data:
            wf.write(word + "\n")
            
    with open(freqs_path, "wb") as ff:
        for _, freq in data:
            # Write frequency as 4-byte little endian unsigned int
            ff.write(struct.pack("<I", freq))
            
    # Compress files
    words_gz = words_path + ".gz"
    freqs_gz = freqs_path + ".gz"
    
    print(f"Compressing to {words_gz}...")
    with open(words_path, "rb") as f_in:
        with gzip.open(words_gz, "wb", compresslevel=9) as f_out:
            f_out.writelines(f_in)
            
    print(f"Compressing to {freqs_gz}...")
    with open(freqs_path, "rb") as f_in:
        with gzip.open(freqs_gz, "wb", compresslevel=9) as f_out:
            f_out.writelines(f_in)
            
    # Clean up uncompressed files
    print("Cleaning up uncompressed files...")
    os.remove(words_path)
    os.remove(freqs_path)
    
    print("Preprocessing completed successfully!")

if __name__ == "__main__":
    preprocess()
